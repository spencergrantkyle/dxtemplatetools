"""
Core Excel flattening functionality.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Optional, Dict, Any
from dataclasses import dataclass
import logging

logger = logging.getLogger(__name__)


@dataclass
class FlattenResult:
    """Result object containing flattened data and metadata."""
    dataframe: pd.DataFrame
    total_cells: int
    sheets_processed: int
    value_types: Dict[str, int]
    file_path: str
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert result to dictionary for JSON serialization."""
        return {
            "total_cells": self.total_cells,
            "sheets_processed": self.sheets_processed,
            "value_types": self.value_types,
            "file_path": self.file_path
        }


def flatten_workbook(input_path: str, include_empty_cells: bool = False) -> FlattenResult:
    """
    Reads an Excel file and flattens all worksheets into a structured dataframe.
    
    Args:
        input_path (str): Path to the Excel workbook (.xlsx or .xlsm)
        include_empty_cells (bool): Whether to include empty cells in output
        
    Returns:
        FlattenResult: Object containing DataFrame and metadata with columns:
            ['Sheet', 'RowNum', 'ColRef', 'Formula Text', 'CellValue', 'Value Type']
            
    Raises:
        FileNotFoundError: If the input file doesn't exist
        ValueError: If the file is not a valid Excel format
        Exception: For other processing errors
    """
    input_file = Path(input_path)
    
    # Validate input file
    if not input_file.exists():
        raise FileNotFoundError(f"Input file '{input_file}' not found.")
    
    if input_file.suffix.lower() not in ['.xlsx', '.xlsm']:
        raise ValueError(f"Input file must be an Excel workbook (.xlsx or .xlsm), got {input_file.suffix}")
    
    logger.info(f"Processing Excel file: {input_file}")
    
    try:
        wb = load_workbook(filename=str(input_file), data_only=False)
        all_data = []
        
        for sheet_name in wb.sheetnames:
            logger.debug(f"Processing sheet: {sheet_name}")
            ws = wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    # Include cell if it has value, is formula, or if including empty cells
                    if cell.value is not None or cell.data_type == 'f' or include_empty_cells:
                        if cell.column is not None:
                            col_ref = get_column_letter(cell.column)
                            row_num = cell.row
                            
                            # Get formula text if it's a formula cell
                            formula_text = cell.value if cell.data_type == 'f' else ''
                            
                            # Get the actual value
                            cell_value = cell.value
                            
                            # Determine value type
                            value_type = _get_value_type(cell)
                            
                            all_data.append({
                                'Sheet': sheet_name,
                                'RowNum': row_num,
                                'ColRef': col_ref,
                                'Formula Text': formula_text,
                                'CellValue': cell_value,
                                'Value Type': value_type
                            })
        
        df = pd.DataFrame(all_data)
        
        # Calculate metadata
        total_cells = len(df)
        sheets_processed = df['Sheet'].nunique() if not df.empty else 0
        value_types = df['Value Type'].value_counts().to_dict() if not df.empty else {}
        
        result = FlattenResult(
            dataframe=df,
            total_cells=total_cells,
            sheets_processed=sheets_processed,
            value_types=value_types,
            file_path=str(input_file)
        )
        
        logger.info(f"Successfully flattened {total_cells} cells from {sheets_processed} sheets")
        return result
        
    except Exception as e:
        logger.error(f"Error processing Excel file {input_file}: {e}")
        raise


def _get_value_type(cell) -> str:
    """Determine the value type of a cell."""
    if cell.data_type == 'f':
        return 'formula'
    elif cell.data_type == 'n':
        return 'number'
    elif cell.data_type == 's':
        return 'string'
    elif cell.data_type == 'b':
        return 'bool'
    elif cell.data_type == 'd':
        return 'date'
    else:
        return 'other'


def save_flattened_data(result: FlattenResult, output_path: Optional[str] = None) -> str:
    """
    Save flattened data to CSV file.
    
    Args:
        result (FlattenResult): The flattening result to save
        output_path (str, optional): Custom output path. If None, generates default name.
        
    Returns:
        str: Path to the saved file
    """
    if output_path is None:
        input_file = Path(result.file_path)
        output_file = input_file.parent / f"{input_file.stem}_flattened.csv"
    else:
        output_file = Path(output_path)
    
    result.dataframe.to_csv(str(output_file), index=False)
    logger.info(f"Saved flattened data to: {output_file}")
    
    return str(output_file)