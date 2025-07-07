import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import argparse
import sys
from pathlib import Path


def flatten_excel_workbook(file_path):
    """
    Flatten an Excel workbook into a DataFrame with metadata for each cell.
    
    Args:
        file_path (str): Path to the Excel workbook
        
    Returns:
        pandas.DataFrame: Flattened data with columns:
            ['Sheet', 'RowNum', 'ColRef', 'Formula Text of cell', 'CellValue', 'Value Type']
    """
    try:
        wb = load_workbook(filename=file_path, data_only=False)
        all_data = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None or cell.data_type == 'f':
                        col_ref = get_column_letter(cell.column)
                        row_num = cell.row
                        
                        # Get formula text if it's a formula cell
                        formula_text = cell.value if cell.data_type == 'f' else ''
                        
                        # Get the actual value (for formulas, this might be None if not calculated)
                        cell_value = cell.value
                        
                        # Determine value type
                        if cell.data_type == 'f':
                            value_type = 'formula'
                        elif cell.data_type == 'n':
                            value_type = 'number'
                        elif cell.data_type == 's':
                            value_type = 'string'
                        elif cell.data_type == 'b':
                            value_type = 'bool'
                        elif cell.data_type == 'd':
                            value_type = 'date'
                        else:
                            value_type = 'other'

                        all_data.append({
                            'Sheet': sheet_name,
                            'RowNum': row_num,
                            'ColRef': col_ref,
                            'Formula Text of cell': formula_text,
                            'CellValue': cell_value,
                            'Value Type': value_type
                        })

        df = pd.DataFrame(all_data)
        return df
        
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        return None


def main():
    parser = argparse.ArgumentParser(description='Flatten Excel workbook into a flat file for LLM processing')
    parser.add_argument('input_file', help='Path to the Excel workbook (.xlsx)')
    parser.add_argument('-o', '--output', help='Output CSV file path (default: input_filename_flattened.csv)')
    
    args = parser.parse_args()
    
    # Validate input file
    input_path = Path(args.input_file)
    if not input_path.exists():
        print(f"Error: Input file '{args.input_file}' not found.")
        sys.exit(1)
    
    if not input_path.suffix.lower() in ['.xlsx', '.xlsm']:
        print(f"Error: Input file must be an Excel workbook (.xlsx or .xlsm)")
        sys.exit(1)
    
    # Determine output file path
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = input_path.parent / f"{input_path.stem}_flattened.csv"
    
    print(f"Processing: {input_path}")
    
    # Flatten the workbook
    flattened_df = flatten_excel_workbook(str(input_path))
    
    if flattened_df is not None:
        # Save to CSV
        flattened_df.to_csv(output_path, index=False)
        print(f"Successfully flattened {len(flattened_df)} cells to: {output_path}")
        
        # Print summary
        print(f"\nSummary:")
        print(f"- Total cells: {len(flattened_df)}")
        print(f"- Sheets processed: {flattened_df['Sheet'].nunique()}")
        print(f"- Value types: {', '.join(flattened_df['Value Type'].value_counts().index.tolist())}")
    else:
        print("Failed to process the Excel file.")
        sys.exit(1)


if __name__ == "__main__":
    main()
