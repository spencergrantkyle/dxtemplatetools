"""
Excel Flattener Tool - Utility functions and data structures.
"""

import pandas as pd
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Optional, Dict, Any
from dataclasses import dataclass
import sys

# Add parent directories to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from shared.excel_io import load_excel_workbook
from shared.logger import get_logger

logger = get_logger(__name__)


@dataclass
class FlattenResult:
    """Result object containing flattened data and metadata."""
    dataframe: pd.DataFrame
    total_cells: int
    sheets_processed: int
    value_types: Dict[str, int]
    input_path: str
    output_path: Optional[str] = None
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert result to dictionary for JSON serialization."""
        return {
            "total_cells": self.total_cells,
            "sheets_processed": self.sheets_processed,
            "value_types": self.value_types,
            "input_path": self.input_path,
            "output_path": self.output_path
        }


def flatten_workbook_to_dataframe(input_path: str, include_empty_cells: bool = False) -> FlattenResult:
    """
    Flatten an Excel workbook into a structured DataFrame.
    
    Args:
        input_path (str): Path to the Excel workbook
        include_empty_cells (bool): Whether to include empty cells
        
    Returns:
        FlattenResult: Object containing DataFrame and metadata
    """
    logger.info(f"Loading workbook: {input_path}")
    wb = load_excel_workbook(input_path, data_only=False)
    
    all_data = []
    
    for sheet_name in wb.sheetnames:
        logger.debug(f"Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        for row in ws.iter_rows():
            for cell in row:
                # Include cell if it has value, is formula, or if including empty cells
                if cell.value is not None or cell.data_type == 'f' or include_empty_cells:
                    if cell.column is not None:
                        col_ref = get_column_letter(cell.column)
                        row_num = cell.row
                        
                        # Get formula text if it's a formula cell
                        formula_text = cell.value if cell.data_type == 'f' else ''
                        
                        # Get the actual value
                        cell_value = cell.value
                        
                        # Determine value type
                        value_type = _get_value_type(cell)
                        
                        all_data.append({
                            'Sheet': sheet_name,
                            'RowNum': row_num,
                            'ColRef': col_ref,
                            'Formula Text': formula_text,
                            'CellValue': cell_value,
                            'Value Type': value_type
                        })
    
    df = pd.DataFrame(all_data)
    
    # Calculate metadata
    total_cells = len(df)
    sheets_processed = df['Sheet'].nunique() if not df.empty else 0
    value_types = df['Value Type'].value_counts().to_dict() if not df.empty else {}
    
    result = FlattenResult(
        dataframe=df,
        total_cells=total_cells,
        sheets_processed=sheets_processed,
        value_types=value_types,
        input_path=str(input_path)
    )
    
    logger.info(f"Successfully flattened {total_cells} cells from {sheets_processed} sheets")
    return result


def _get_value_type(cell) -> str:
    """Determine the value type of a cell."""
    if cell.data_type == 'f':
        return 'formula'
    elif cell.data_type == 'n':
        return 'number'
    elif cell.data_type == 's':
        return 'string'
    elif cell.data_type == 'b':
        return 'bool'
    elif cell.data_type == 'd':
        return 'date'
    else:
        return 'other'


def print_summary(result: FlattenResult) -> None:
    """Print a formatted summary of the flattening result."""
    print(f"\n📊 Flattening Summary:")
    print(f"   📁 File: {result.input_path}")
    print(f"   📋 Total cells: {result.total_cells:,}")
    print(f"   📄 Sheets processed: {result.sheets_processed}")
    
    if result.value_types:
        print(f"   🏷️  Value types:")
        for value_type, count in result.value_types.items():
            print(f"      • {value_type}: {count:,}")


def create_analysis_preview(result: FlattenResult, max_rows: int = 10) -> str:
    """
    Create a preview of the flattened data for analysis.
    
    Args:
        result (FlattenResult): The flattening result
        max_rows (int): Maximum number of rows to include in preview
        
    Returns:
        str: Formatted preview text
    """
    df = result.dataframe
    
    preview_text = f"""
Excel Workbook Analysis Preview
==============================

File: {result.input_path}
Total Cells: {result.total_cells:,}
Sheets: {result.sheets_processed}
Value Types: {', '.join(f'{k}({v})' for k, v in result.value_types.items())}

Sample Data (first {min(max_rows, len(df))} rows):
{df.head(max_rows).to_string(index=False)}

Sheet Distribution:
{df['Sheet'].value_counts().to_string()}
"""
    
    return preview_text