"""
Shared Excel I/O operations.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Optional, Dict, Any, List
import logging

logger = logging.getLogger(__name__)


def load_excel_workbook(file_path: str, data_only: bool = False):
    """
    Load an Excel workbook using openpyxl.
    
    Args:
        file_path (str): Path to the Excel file
        data_only (bool): If True, only cell values are loaded (no formulas)
        
    Returns:
        openpyxl.Workbook: Loaded workbook object
        
    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If file is not valid Excel format
    """
    file_path_obj = Path(file_path)
    
    if not file_path_obj.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path_obj}")
    
    if file_path_obj.suffix.lower() not in ['.xlsx', '.xlsm']:
        raise ValueError(f"File must be Excel format (.xlsx/.xlsm), got: {file_path_obj.suffix}")
    
    logger.info(f"Loading Excel workbook: {file_path_obj}")
    return load_workbook(filename=str(file_path_obj), data_only=data_only)


def save_to_csv(dataframe: pd.DataFrame, output_path: str, **kwargs) -> str:
    """
    Save DataFrame to CSV file.
    
    Args:
        dataframe (pd.DataFrame): Data to save
        output_path (str): Output file path
        **kwargs: Additional arguments for pandas.to_csv()
        
    Returns:
        str: Path to saved file
    """
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    # Default CSV options
    csv_options = {
        'index': False,
        'encoding': 'utf-8'
    }
    csv_options.update(kwargs)
    
    dataframe.to_csv(str(output_file), **csv_options)
    logger.info(f"Saved {len(dataframe)} rows to: {output_file}")
    
    return str(output_file)


def find_guid_in_column_a(worksheet, guid: str) -> Optional[int]:
    """
    Find a GUID in column A of a worksheet.
    
    Args:
        worksheet: openpyxl worksheet object
        guid (str): 8-character GUID to search for
        
    Returns:
        Optional[int]: Row number if found, None otherwise
    """
    try:
        for row in range(1, worksheet.max_row + 1):
            cell_value = worksheet[f'A{row}'].value
            if cell_value and str(cell_value).strip() == guid.strip():
                logger.debug(f"Found GUID {guid} at row {row}")
                return row
        
        logger.warning(f"GUID {guid} not found in column A")
        return None
        
    except Exception as e:
        logger.error(f"Error searching for GUID {guid}: {e}")
        return None


def extract_row_data(worksheet, row_num: int, start_col: int = 124, end_col: int = 200) -> List[Any]:
    """
    Extract data from specific columns in a row.
    
    Args:
        worksheet: openpyxl worksheet object
        row_num (int): Row number to extract from
        start_col (int): Starting column index (1-based)
        end_col (int): Ending column index (1-based, inclusive)
        
    Returns:
        List[Any]: Values from the specified columns
    """
    data = []
    for col in range(start_col, end_col + 1):
        try:
            value = worksheet.cell(row=row_num, column=col).value
            data.append(value)
        except Exception as e:
            logger.warning(f"Error reading cell ({row_num}, {col}): {e}")
            data.append(None)
    
    return data


def get_column_headers(start_col: int = 124, end_col: int = 200) -> List[str]:
    """
    Generate column headers for the specified range.
    
    Args:
        start_col (int): Starting column index
        end_col (int): Ending column index (inclusive)
        
    Returns:
        List[str]: Column headers like ['Col124', 'Col125', ...]
    """
    return [f'Col{i}' for i in range(start_col, end_col + 1)]